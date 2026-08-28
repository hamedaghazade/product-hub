from io import BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenPyXLImage
from app.services.barcode_service import BarcodeService

class ExcelExportService:
    @staticmethod
    def generate_product_catalog(products: List[Dict[str, Any]]) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "کاتالوگ محصولات"
        ws.sheet_view.rightToLeft = True

        # استایل‌بندی هدرها
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
        cell_font = Font(name="Tahoma", size=9)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        headers = ["ردیف", "نام کالا", "قیمت خرید (تومان)", "تعداد در بسته", "قیمت مصرف‌کننده (تومان)", "کد بارکد", "تصویر بارکد"]
        ws.append(headers)

        # تنظیم ارتفاع و رنگ هدر
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # عرض ستون‌ها
        column_widths = {'A': 8, 'B': 30, 'C': 18, 'D': 14, 'E': 20, 'F': 18, 'G': 28}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # درج داده‌ها و بارکدها
        for idx, item in enumerate(products, start=1):
            row_num = idx + 1
            ws.row_dimensions[row_num].height = 70  # فضای کافی برای تصویر بارکد

            ws.cell(row=row_num, column=1, value=idx).alignment = align_center
            ws.cell(row=row_num, column=2, value=item.get("title", "")).alignment = Alignment(horizontal="right", vertical="center")
            
            c_price = ws.cell(row=row_num, column=3, value=item.get("cost_price", 0))
            c_price.number_format = '#,##0'
            c_price.alignment = align_center

            ws.cell(row=row_num, column=4, value=item.get("units_per_pack", 1)).alignment = align_center

            cons_price = ws.cell(row=row_num, column=5, value=item.get("consumer_price", 0))
            cons_price.number_format = '#,##0'
            cons_price.alignment = align_center

            ws.cell(row=row_num, column=6, value=str(item.get("barcode_value", ""))).alignment = align_center

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).font = cell_font
                ws.cell(row=row_num, column=c).border = thin_border

            # ساخت و درج تصویر بارکد در ستون G
            barcode_str = str(item.get("barcode_value", ""))
            if barcode_str:
                barcode_img_stream = BarcodeService.generate_barcode_image(
                    title=item.get("title", ""),
                    barcode_value=barcode_str
                )
                img = OpenPyXLImage(barcode_img_stream)
                img.width = 170
                img.height = 80
                cell_address = f"G{row_num}"
                ws.add_image(img, cell_address)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output