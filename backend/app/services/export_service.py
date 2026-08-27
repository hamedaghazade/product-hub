import io
import logging
from typing import List, Dict, Any
import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
import arabic_reshaper
from bidi.algorithm import get_display

from app.services.barcode_service import BarcodeService

logger = logging.getLogger(__name__)

class ExportService:
    @staticmethod
    def _prepare_persian_text(text: Any) -> str:
        """اصلاح اتصالات و راست‌به‌چپ کردن متون فارسی برای موتورهای غیروب"""
        if text is None:
            return ""
        text_str = str(text).strip()
        if not text_str:
            return ""
        try:
            reshaped = arabic_reshaper.reshape(text_str)
            return get_display(reshaped)
        except Exception:
            return text_str

    @classmethod
    def export_excel(cls, products: List[Dict[str, Any]]) -> io.BytesIO:
        """تولید فایل اکسل راست‌چین همراه با تصویر بارکد در سلول‌ها"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "لیست محصولات"
        ws.views.sheetView[0].rightToLeft = True

        headers = [
            "ردیف",
            "نام محصول",
            "تعداد در بسته",
            "قیمت خرید (ریال)",
            "کد بارکد",
            "تصویر بارکد",
            "قیمت مصرف‌کننده (ریال)"
        ]
        ws.append(headers)

        # استایل‌های هدر
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        data_font = Font(name="Tahoma", size=9)

        for col_idx in range(1, 8):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30

        for idx, item in enumerate(products, start=2):
            ws.row_dimensions[idx].height = 55

            # ۱. ردیف
            c1 = ws.cell(row=idx, column=1, value=idx - 1)
            c1.alignment = Alignment(horizontal="center", vertical="center")
            
            # ۲. نام محصول
            c2 = ws.cell(row=idx, column=2, value=str(item.get("title", "")))
            c2.alignment = Alignment(horizontal="right", vertical="center")

            # ۳. تعداد در بسته
            c3 = ws.cell(row=idx, column=3, value=item.get("units_per_pack", 1))
            c3.alignment = Alignment(horizontal="center", vertical="center")

            # ۴. قیمت خرید
            c4 = ws.cell(row=idx, column=4, value=item.get("cost_price", 0))
            c4.number_format = '#,##0'
            c4.alignment = Alignment(horizontal="center", vertical="center")

            # ۵. کد بارکد
            c5 = ws.cell(row=idx, column=5, value=str(item.get("barcode_value", "")))
            c5.alignment = Alignment(horizontal="center", vertical="center")

            # ۶. درج تصویر بارکد
            barcode_val = str(item.get("barcode_value", "")).strip()
            item_title = str(item.get("title", "")).strip()
            if barcode_val:
                try:
                    img_stream = BarcodeService.generate_barcode(code=barcode_val, title=item_title)
                    xl_img = OpenPyxlImage(img_stream)
                    xl_img.width = 115
                    xl_img.height = 48
                    ws.add_image(xl_img, f"F{idx}")
                except Exception as e:
                    logger.warning(f"عدم امکان افزودن تصویر بارکد برای ردیف {idx}: {str(e)}")
                    ws.cell(row=idx, column=6, value="—").alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws.cell(row=idx, column=6, value="—").alignment = Alignment(horizontal="center", vertical="center")

            # ۷. قیمت مصرف کننده
            c7_val = item.get("consumer_price")
            c7 = ws.cell(row=idx, column=7, value=c7_val if c7_val is not None else "—")
            if isinstance(c7_val, (int, float)):
                c7.number_format = '#,##0'
            c7.alignment = Alignment(horizontal="center", vertical="center")

            # اعمال فونت و بوردر برای کل خانه‌های ردیف
            for col_idx in range(1, 8):
                cell = ws.cell(row=idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

        # تنظیم پهنای استاندارد ستون‌ها
        col_widths = {1: 8, 2: 32, 3: 15, 4: 18, 5: 18, 6: 22, 7: 20}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def export_pdf(cls, products: List[Dict[str, Any]]) -> io.BytesIO:
        """تولید خروجی استاندارد و آماده چاپ PDF با پشتیبانی از یونیکد فارسی"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=25,
            bottomMargin=25
        )
        elements = []

        title_style = ParagraphStyle(
            name='HeaderStyle',
            fontName='Helvetica-Bold',
            fontSize=15,
            leading=18,
            alignment=1,
            textColor=colors.HexColor("#0F172A")
        )
        elements.append(Paragraph(cls._prepare_persian_text("کاتالوگ و لیست قیمت محصولات"), title_style))
        elements.append(Spacer(1, 15))

        # ساختار ستون‌های جدول راست به چپ
        table_headers = [
            cls._prepare_persian_text("ردیف"),
            cls._prepare_persian_text("نام کالا"),
            cls._prepare_persian_text("بسته"),
            cls._prepare_persian_text("قیمت خرید"),
            cls._prepare_persian_text("کد بارکد"),
            cls._prepare_persian_text("تصویر بارکد"),
            cls._prepare_persian_text("قیمت مصرف")
        ]
        table_data = [table_headers]

        for idx, p in enumerate(products, start=1):
            reshaped_title = cls._prepare_persian_text(p.get("title", ""))
            barcode_val = str(p.get("barcode_value", "")).strip()

            img_element = "—"
            if barcode_val:
                try:
                    barcode_stream = BarcodeService.generate_barcode(code=barcode_val, title="")
                    img_element = ReportLabImage(barcode_stream, width=80, height=28)
                except Exception:
                    pass

            cost_price_val = f"{p.get('cost_price', 0):,}"
            consumer_price_val = f"{p.get('consumer_price', 0):,}" if p.get('consumer_price') else "—"

            row = [
                str(idx),
                reshaped_title,
                str(p.get("units_per_pack", 1)),
                cost_price_val,
                barcode_val,
                img_element,
                consumer_price_val
            ]
            table_data.append(row)

        pdf_table = Table(table_data, colWidths=[25, 125, 35, 75, 80, 95, 75])
        pdf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))

        elements.append(pdf_table)
        doc.build(elements)
        buffer.seek(0)
        return buffer